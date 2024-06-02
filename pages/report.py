import openpyxl
import os
from datetime import datetime
from flet import *
from openpyxl.styles import Font, Alignment, Border, Side
from service.connection import *
from utils.consts import *

class Report(Container):
    def __init__(self, page: Page):
        super().__init__()
        self.page = page
        self.page.theme_mode = ThemeMode.LIGHT
        self.alignment = alignment.center
        self.expand = True
        self.bgcolor = primary_colors['WHITE']
        self.user_id = self.page.session.get('user_id')

        self.dropdown_options_specialists = []
        self.dropdown_options_departments = []

        # * DROPDOWN LISTS
        self.report_template = Container(
            content=Dropdown(
                # label='Специалист',
                hint_text='Выберите шаблон',
                color='black',
                # content_padding=30,
                options=[
                    dropdown.Option('Карта КПЭ'),
                    dropdown.Option('Расчет премии'),
                    dropdown.Option('Сводные данные по исполнению')
                ],
                on_change=self.show_department
            )
        )

        self.report_spec = Container(
            content=Dropdown(
                # label='Специалист',
                hint_text='Выберите специалиста',
                color='black',
                # content_padding=30,
                options=self.dropdown_options_specialists
            )
        )

        self.report_depart = Container(
            content=Dropdown(
                # label='Специалист',
                hint_text='Выберите управление',
                color='black',
                # content_padding=30,
                options=self.dropdown_options_departments,
                on_change=self.show_specialists
            )
        )

        self.report_quater = Container(
            content=Dropdown(
                # label='Специалист',
                hint_text='Выберите квартал',
                color='black',
                # content_padding=30,
                options=[
                    dropdown.Option('1 квартал'),
                    dropdown.Option('2 квартал'),
                    dropdown.Option('3 квартал'),
                    dropdown.Option('4 квартал'),
                ]
            )
        )
        # *TEXT
        self.number_of_version_kpe = Container(
            Text(
                value=""
            )
        )

        # *ALTER DIALOG
        self.alter_dialog = AlertDialog(
            modal=True,
            title=Text("Формирование отчета"),
            content=Column(
                height=250,
                width=700,
                controls=[
                    self.report_depart,
                    self.report_spec,
                    self.report_quater
                ]
            ),
            actions=[
                TextButton("Сформировать", on_click=self.alter_dialog_select_columns_data),
                TextButton("Закрыть", on_click=self.close_dlg),
            ],
            actions_alignment=MainAxisAlignment.END,
            on_dismiss=lambda e: print("Modal dialog dismissed!"),
        )
        self.alter_dialog_kpe = AlertDialog(
            modal=True,
            title=Text("Формирование отчета"),
            content=Column(
                height=250,
                width=700,
                controls=[
                    self.report_depart,
                    self.report_spec
                ]
            ),
            actions=[
                TextButton("Сформировать", on_click=self.select_kpe),
                TextButton("Закрыть", on_click=self.close_dlg_kpe),
            ],
            actions_alignment=MainAxisAlignment.END,
            on_dismiss=lambda e: print("Modal dialog dismissed!"),
        )
        self.alter_dialog_summary = AlertDialog(
            modal=True,

            title=Text("Формирование отчета"),
            content=Column(
                height=250,
                width=700,
                controls=[
                    self.report_depart,
                    self.report_quater
                ]
            ),
            actions=[
                TextButton("Сформировать", on_click=self.select_summary),
                TextButton("Закрыть", on_click=self.close_dlg_summary),
            ],
            actions_alignment=MainAxisAlignment.END,
            on_dismiss=lambda e: print("Modal dialog dismissed!"),
        )
        self.alter_dialog_block = AlertDialog(
            modal=True,
            title=Text("Default Title"),
            content=Text("Default Content"),
            actions=[TextButton("OK", on_click=self.close_dlg_ok)],
            on_dismiss=lambda e: print("Modal dialog dismissed!")
        )

        # !UI
        self.content = ListView(
            spacing=0,
            controls=[
                Container(
                    width=8000,
                    padding=40,
                    bgcolor=primary_colors['GREEN'],
                    # border_radius=15,
                    content=Column(
                        horizontal_alignment='center',  # Align the text to the right
                        controls=[
                            Container(
                                # alignment='center',
                                content=Row(
                                    alignment='spaceBetween',
                                    controls=[
                                        Container(
                                            # width=200,
                                            bgcolor=primary_colors['WHITE'],
                                            width=70,
                                            height=70,
                                            border_radius=50,
                                            content=IconButton(
                                                icons.ARROW_BACK_OUTLINED,
                                                icon_color=primary_colors['GREEN'],
                                                icon_size=30,
                                                on_click=lambda x: x == self.page.go('/home')
                                            )
                                        ),
                                        Container(
                                            content=Text(
                                                value='Отчеты',
                                                size=18,
                                                color=primary_colors['WHITE'],
                                                text_align='center',
                                            ),
                                        ),
                                        Container(width=70),
                                    ]
                                )
                            )
                        ],
                    )
                ),

                # manual buttons
                Container(
                    expand=True,
                    bgcolor=primary_colors['WHITE'],
                    content=Column(
                        expand=True,
                        # alignment='center',
                        horizontal_alignment='center',
                        controls=[
                            # 1st row
                            Container(height=50),
                            Container(
                                Row(
                                    spacing='50',
                                    alignment='center',
                                    # horizontal_alignment='center',
                                    controls=[
                                        self.report_template,
                                        # self.report_spec,
                                        # self.report_quater,
                                        ElevatedButton(
                                            color=primary_colors['WHITE'],
                                            bgcolor=primary_colors['GREEN'],
                                            width=250,
                                            height=70,
                                            content=Column(
                                                horizontal_alignment='center',
                                                alignment='center',
                                                controls=[
                                                    Container(
                                                        Text(
                                                            value='Сформировать',
                                                            size=16,
                                                            color=primary_colors['WHITE'],
                                                            text_align='center',
                                                            weight='bold',
                                                        )
                                                    )
                                                ]
                                            ),
                                            on_click=self.form_report
                                            # on_click=lambda x: x == self.page.go('/home')
                                        ),
                                        ElevatedButton(
                                            color=primary_colors['WHITE'],
                                            bgcolor=primary_colors['GREEN'],
                                            width=250,
                                            height=70,
                                            content=Column(
                                                horizontal_alignment='center',
                                                alignment='center',
                                                controls=[
                                                    Container(
                                                        Text(
                                                            value='Экспорт',
                                                            size=16,
                                                            color=primary_colors['WHITE'],
                                                            text_align='center',
                                                            weight='bold',
                                                        )
                                                    )
                                                ]
                                            ),
                                            # on_click=lambda _: self.file_picker.pick_files()
                                            on_click=self.export_report_to_excel
                                        )
                                    ]
                                )
                            ),

                            # Container(height=50),
                            # 2nd row
                            Container(
                                content=DataTable(
                                    columns=[],
                                    rows=[],  # Leave this empty for now
                                    border=border.all(1, primary_colors['BLACK']),
                                    vertical_lines=border.BorderSide(1, primary_colors['BLACK']),
                                    horizontal_lines=border.BorderSide(1, primary_colors['BLACK']),
                                    sort_column_index=0,
                                    sort_ascending=True,
                                    heading_row_color=colors.BLACK12,
                                    heading_row_height=100,
                                    data_row_max_height=80,
                                    divider_thickness=0,

                                ),
                                alignment=alignment.center,
                                padding=padding.all(20),
                            ),
                            Container(
                                content=Row(
                                    # width=1000,
                                    controls=[
                                        Container(width=800),
                                        Container(width=700),
                                        Container(content=self.number_of_version_kpe)
                                    ]
                                )
                            )
                        ],
                    )
                ),
            ]
        )
    def show_department(self, e):
        self.dropdown_options_departments.clear()
        try:
            cursor = connection.cursor()
            cursor.execute('SELECT name FROM name_of_department ORDER BY department_id')
            results = cursor.fetchall()

            for row in results:
                self.dropdown_options_departments.append(dropdown.Option(row[0]))
            
            if self.report_template.content.value == "Сводные данные по исполнению":
                self.dropdown_options_departments.append(dropdown.Option("Все управления"))
            else:
                pass
        except Exception as e:
            print(f"Error fetching data from the database: {str(e)}")


    def show_specialists(self, e):
        self.dropdown_options_specialists.clear()
        try:
            cursor = connection.cursor()
            cursor.execute(f"SELECT department_id FROM name_of_department WHERE name = '{self.report_depart.content.value}'")
            specialist_department_id = cursor.fetchone()[0]
            
            cursor.execute(f"SELECT full_name FROM specialists WHERE specialist_department_id = {specialist_department_id}")
            specialists_full_name = cursor.fetchall()

            for row in specialists_full_name:
                self.dropdown_options_specialists.append(dropdown.Option(row[0]))
                
            self.page.update()
        except Exception as e:
            print(f"Error fetching data from the database: {str(e)}")
    
    def show_block_dialog(self, content_text, title_text):
        self.page.dialog = self.alter_dialog_block
        self.alter_dialog_block.content = Text(f"{content_text}")
        self.alter_dialog_block.title = Text(f"{title_text}")
        self.alter_dialog_block.open = True
        self.page.update()

    def close_dlg_ok(self, e):
        self.page.dialog = self.alter_dialog_block
        self.alter_dialog_block.open = False
        print("Вы закрыли модульное окно успеха")
        self.page.update()

    # TODO: FOR Premium calculation
    def open_dialog(self):
        self.page.dialog = self.alter_dialog
        self.alter_dialog.open = True
        self.page.update()

#! PREMIUM CALCULATION REPORT
    def alter_dialog_select_columns_data(self, e):
        try:
            cursor = connection.cursor()
            # Получение ID специалиста
            cursor.execute(
                f"SELECT specialist_id FROM specialists WHERE full_name='{str(self.report_spec.content.value)}';")
            specialist_id = cursor.fetchone()[0]
            cursor.execute(
                    f"""
                    SELECT 
                        concat(
                            toString(kpe_user_id), '-',
                            substring(replace(toString(date), '-', ''), 7, 2),
                            substring(replace(toString(date), '-', ''), 5, 2),
                            substring(replace(toString(date), '-', ''), 1, 4),
                            '-', toString(number)
                        ) AS number_of_version
                    FROM kpe_table
                    WHERE 
                        kpe_specialist_id = {int(specialist_id)} AND 
                        date = (
                            SELECT MAX(date)
                            FROM kpe_table
                            WHERE kpe_specialist_id = {int(specialist_id)}
                        ) AND 
                        number = (
                            SELECT MAX(number)
                            FROM kpe_table
                            WHERE 
                                kpe_specialist_id = {int(specialist_id)} AND 
                                date = (
                                    SELECT MAX(date)
                                    FROM kpe_table
                                    WHERE kpe_specialist_id = {int(specialist_id)}
                                )
                        );
                    """
            )
            result = cursor.fetchone()
            latest_version = result[0] if result is not None else '0'

            # Получение максимальных значений date и number для kpe_table
            cursor.execute(
                f"""
                SELECT MAX(date), MAX(number)
                FROM kpe_table
                WHERE kpe_specialist_id = {specialist_id}
                """
            )
            kpe_max_date, kpe_max_number = cursor.fetchone()

            # Получение максимальных значений date и number для actual_value
            cursor.execute(
                f"""
                SELECT MAX(date), MAX(number)
                FROM actual_value
                WHERE actual_specialist_id = {specialist_id}
                """
            )
            actual_max_date, actual_max_number = cursor.fetchone()

            print(actual_max_date, kpe_max_date, actual_max_number, kpe_max_number)

            # Определение колонок для квартала
            quarter_mapping = {
                "1 квартал": ("`1st_quater_value`", "KPE_weight_1", "1-й квартал"),
                "2 квартал": ("`2nd_quater_value`", "KPE_weight_2", "2-й квартал"),
                "3 квартал": ("`3rd_quater_value`", "KPE_weight_3", "3-й квартал"),
                "4 квартал": ("`4th_quater_value`", "KPE_weight_4", "4-й квартал")
            }
            quater_column, weight_column, actual_quater_value = quarter_mapping.get(self.report_quater.content.value, (None, None, None))

            if not quater_column:
                self.show_block_dialog("Вы не выбрали номер квартала", "Ошибка")
                return

            # Формирование запроса
            query = f"""
            SELECT
                ROW_NUMBER() OVER (ORDER BY kpe_id) AS "порядковый номер",
                ni.name AS indicator_name,
                um.type AS unit_of_measurement,
                {quater_column},
                av.value AS actual_value,
                {weight_column},
                multiIf({quater_column} <= av.value, {weight_column} / 100, 0) AS bonus_share
            FROM kpe_table AS kt
            INNER JOIN actual_value AS av ON kt.kpe_indicators_id = av.actual_indicators_id
            INNER JOIN name_of_indicators AS ni ON kt.kpe_indicators_id = ni.indicators_id
            INNER JOIN units_of_measurement AS um ON kt.kpe_units_id = um.measurement_id
            WHERE kt.kpe_specialist_id = {specialist_id}
            AND kt.date = '{kpe_max_date}'
            AND kt.number = {kpe_max_number}
            AND av.date = '{actual_max_date}'
            AND av.number = {actual_max_number}
            AND av.actual_specialist_id = kt.kpe_specialist_id
            AND av.actual_indicators_id = kt.kpe_indicators_id
            AND av.quarter_number = '{actual_quater_value}'
            AND kt.status = 'Активно'
            ORDER BY kt.kpe_id ASC
            """
            print(query)

            # Определите структуру колонок для "Премии"
            columns = [
                DataColumn(Text("№ пп."), numeric=True),
                DataColumn(Text("Наименование показателя")),
                DataColumn(Text("Ед.изм.")),
                DataColumn(Text("План")),
                DataColumn(Text("Факт")),
                DataColumn(Text("Вес КПЭ, %")),
                DataColumn(Text("Доля премии по факту выполнения показателя"))
            ]

            cursor.execute(query)
            global results_premi
            results_premi = cursor.fetchall()

            query_result = results_premi

            data_rows = []

            for row in query_result:
                cells = [DataCell(Text(str(value))) for value in row]
                data_row = DataRow(cells=cells)
                data_rows.append(data_row)
            self.content.controls[1].content.controls[2].content.columns = columns
            self.content.controls[1].content.controls[2].content.rows = data_rows

            self.page.dialog = self.alter_dialog
            self.alter_dialog.open = False
            self.number_of_version_kpe.content.value = latest_version
            self.page.update()
        except Exception as e:
            self.show_block_dialog("Вы не выбрали специалиста или номер квартала", "Ошибка")
            print(f"Ошибка: {str(e)}")

    def close_dlg(self, e):
        self.page.dialog = self.alter_dialog
        self.alter_dialog.open = False
        self.page.update()

    # TODO: FOR KPE TABLE
    def open_dialog_kpe(self):
        self.page.dialog = self.alter_dialog_kpe
        self.alter_dialog_kpe.open = True
        self.page.update()

#! KPE MAP REPORT
    def select_kpe(self, e):
        try:
            cursor = connection.cursor()
            cursor.execute(
                f"SELECT specialist_id FROM specialists WHERE full_name='{str(self.report_spec.content.value)}';")
            specialist_id = cursor.fetchone()[0]

            cursor.execute(
                        f"""
                        SELECT 
                            concat(
                                toString(kpe_user_id), '-',
                                substring(replace(toString(date), '-', ''), 7, 2),
                                substring(replace(toString(date), '-', ''), 5, 2),
                                substring(replace(toString(date), '-', ''), 1, 4),
                                '-', toString(number)
                            ) AS number_of_version
                        FROM kpe_table
                        WHERE 
                            kpe_specialist_id = {int(specialist_id)} AND 
                            date = (
                                SELECT MAX(date)
                                FROM kpe_table
                                WHERE kpe_specialist_id = {int(specialist_id)}
                            ) AND 
                            number = (
                                SELECT MAX(number)
                                FROM kpe_table
                                WHERE 
                                    kpe_specialist_id = {int(specialist_id)} AND 
                                    date = (
                                        SELECT MAX(date)
                                        FROM kpe_table
                                        WHERE kpe_specialist_id = {int(specialist_id)}
                                    )
                            );
                        """
            )
            result = cursor.fetchone()
            latest_version = result[0] if result is not None else '0'
            print(specialist_id)
            print(latest_version)

            query = f"""
                SELECT
                    ROW_NUMBER() OVER (ORDER BY kpe_id) AS "порядковый номер",
                    ni.name AS indicator_name,
                    um.type AS unit_of_measurement,
                    kt.1st_quater_value,
                    kt.2nd_quater_value,
                    kt.3rd_quater_value,
                    kt.4th_quater_value,
                    kt.year,
                    kt.KPE_weight_1,
                    kt.KPE_weight_2,
                    kt.KPE_weight_3,
                    kt.KPE_weight_4
                FROM kpe_table AS kt
                JOIN name_of_indicators AS ni ON kt.kpe_indicators_id = ni.indicators_id
                JOIN units_of_measurement AS um ON kt.kpe_units_id = um.measurement_id
                WHERE
                    kt.kpe_specialist_id = {int(specialist_id)} AND
                    kt.status = 'Активно' AND
                    kt.date = (
                        SELECT MAX(date)
                        FROM kpe_table
                        WHERE kpe_specialist_id = {int(specialist_id)} AND status = 'Активно'
                    ) AND
                    kt.number = (
                        SELECT MAX(number)
                        FROM kpe_table
                        WHERE
                            kpe_specialist_id = {int(specialist_id)} AND
                            status = 'Активно' AND
                            date = (
                                SELECT MAX(date)
                                FROM kpe_table
                                WHERE kpe_specialist_id = {int(specialist_id)} AND status = 'Активно'
                            )
                    )
                ORDER BY kpe_id;
            """
            print(query)

            columns = [
                DataColumn(Text("п/п"), numeric=True),
                DataColumn(Text("Наименование показателя")),
                DataColumn(Text("Ед.изм.")),
                DataColumn(Text("""1
    кв."""), numeric=True),
                DataColumn(Text("""2
    кв."""), numeric=True),
                DataColumn(Text("""3
    кв."""), numeric=True),
                DataColumn(Text("""4
    кв."""), numeric=True),
                DataColumn(Text("год")),
                DataColumn(Text("""   Вес КПЭ 
    1 кв."""), numeric=True),
                DataColumn(Text("""   Вес КПЭ 
    2 кв."""), numeric=True),
                DataColumn(Text("""   Вес КПЭ 
    3 кв."""), numeric=True),
                DataColumn(Text("""   Вес КПЭ 
    4 кв."""), numeric=True),
            ]
            self.number_of_version_kpe.content.value = latest_version

            cursor.execute(query)
            global results
            results = cursor.fetchall()
            query_result = results

            data_rows = []

            for row in query_result:
                cells = [DataCell(Text(str(value))) for value in row]
                data_row = DataRow(cells=cells)
                data_rows.append(data_row)
            self.content.controls[1].content.controls[2].content.columns = columns
            self.content.controls[1].content.controls[2].content.rows = data_rows
            self.page.dialog = self.alter_dialog_kpe
            self.alter_dialog_kpe.open = False
            self.page.update()
        except TypeError:
            self.show_block_dialog("Карта КПЭ специалиста не заполнена", "Ошибка")

    def close_dlg_kpe(self, e):
        self.page.dialog = self.alter_dialog_kpe
        self.alter_dialog_kpe.open = False
        self.page.update()

#! SUMMARY REPORT
    def open_dialog_summary(self):
        self.page.dialog = self.alter_dialog_summary
        self.alter_dialog_summary.open = True
        self.page.update()

    def select_summary(self, e):
        try:
            cursor = connection.cursor()
            query_select = f'SELECT MAX(date), MAX(number) FROM kpe_table;'
            cursor.execute(query_select)
            latest_date, latest_number = cursor.fetchone()

            query_select_actual = f'SELECT MAX(date) FROM actual_value;'
            cursor.execute(query_select_actual)
            latest_date_actual = cursor.fetchone()[0]

            quater = self.report_quater.content.value

            if quater == "1 квартал":
                quater_column = "1st_quater_value"
                weight_column = "KPE_weight_1"
                actual_quater_value = "1-й квартал"
            elif quater == "2 квартал":
                quater_column = "2nd_quater_value"
                weight_column = "KPE_weight_2"
                actual_quater_value = "2-й квартал"
            elif quater == "3 квартал":
                quater_column = "3rd_quater_value"
                weight_column = "KPE_weight_3"
                actual_quater_value = "3-й квартал"
            else:
                quater_column = "4th_quater_value"
                weight_column = "KPE_weight_4"
                actual_quater_value = "4-й квартал"

            if self.report_depart.content.value == "Все управления":
                query = f"""
                    SELECT
                        ROW_NUMBER() OVER () AS `порядковый номер`,
                        'агентство по труду и занятости населения Сахалинской области' AS `Наименование структурного подразделения Правительства Сахалинской области, государственного органа или органа исполнительной власти Сахалинской области`,
                        nd.name AS `Наименование структурного подразделения органа исполнительной власти Сахалинской области`,
                        s.position AS `Должность`,
                        s.full_name AS `ФИО`,
                        SUM( (kt.`{weight_column}` / 100) * ( (av.value * 100) / NULLIF(kt.`{quater_column}`, 0)
                            )
                        ) AS `Процент выполнения`
                    FROM kpe_table AS kt
                        INNER JOIN specialists AS s ON kt.kpe_specialist_id = s.specialist_id
                        INNER JOIN name_of_department AS nd ON s.specialist_department_id = nd.department_id
                        INNER JOIN actual_value AS av ON kt.kpe_specialist_id = av.actual_specialist_id AND kt.kpe_indicators_id = av.actual_indicators_id
                    WHERE
                        av.quarter_number = '{actual_quater_value}'
                        AND kt.status = 'Активно'
                        AND toYear(kt.date) = toYear(now())
                        AND toYear(av.date) = toYear(now())
                    GROUP BY
                        `Наименование структурного подразделения Правительства Сахалинской области, государственного органа или органа исполнительной власти Сахалинской области`,
                        `Наименование структурного подразделения органа исполнительной власти Сахалинской области`,
                        `Должность`,
                        `ФИО`
                """
                print(query)
            else:
                query_select_dep_id = "SELECT department_id FROM name_of_department WHERE name = '{}'".format(
                    self.report_depart.content.value)
                cursor.execute(query_select_dep_id)
                department_id = cursor.fetchone()[0]

                query = f"""
                    SELECT
                        ROW_NUMBER() OVER () AS `порядковый номер`,
                        'агентство по труду и занятости населения Сахалинской области' AS `Наименование структурного подразделения Правительства Сахалинской области, государственного органа или органа исполнительной власти Сахалинской области`,
                        nd.name AS `Наименование структурного подразделения органа исполнительной власти Сахалинской области`,
                        s.position AS `Должность`,
                        s.full_name AS `ФИО`,
                        SUM( (kt.`{weight_column}` / 100) * ( (av.value * 100) / NULLIF(kt.`{quater_column}`, 0)
                            )
                        ) AS `Процент выполнения`
                    FROM kpe_table AS kt
                        INNER JOIN specialists AS s ON kt.kpe_specialist_id = s.specialist_id
                        INNER JOIN name_of_department AS nd ON s.specialist_department_id = nd.department_id
                        INNER JOIN actual_value AS av ON kt.kpe_specialist_id = av.actual_specialist_id AND kt.kpe_indicators_id = av.actual_indicators_id
                    WHERE
                        s.specialist_department_id = {department_id}
                        AND av.quarter_number = '{actual_quater_value}'
                        AND kt.status = 'Активно'
                        AND toYear(kt.date) = toYear(now())
                        AND toYear(av.date) = toYear(now())
                    GROUP BY
                        `Наименование структурного подразделения Правительства Сахалинской области, государственного органа или органа исполнительной власти Сахалинской области`,
                        `Наименование структурного подразделения органа исполнительной власти Сахалинской области`,
                        `Должность`,
                        `ФИО`
                """

            print(query)
            columns = [
                DataColumn(Text("""№
    п/п"""), numeric=True),
                DataColumn(Text("""Наименование структурного
    подразделения Правительства
    Сахалинской области,
    государственного органа
    или органа исполнительной
    власти Сахалинской области""")),
                DataColumn(Text("""Наименование структурного
    подразделения органа
    исполнительной власти
    Сахалинской области""")),
                DataColumn(Text("Должность")),
                DataColumn(Text("ФИО")),
                DataColumn(Text("""Процент выполнения
    КПЭ по итогам
    отчетного периода, 
    %""")),
            ]

            cursor.execute(query)
            global results_summary
            results_summary = cursor.fetchall()
            print(results_summary)

            query_result = results_summary

            data_rows = []

            for row in query_result:
                cells = [DataCell(Text(str(value))) for value in row]
                data_row = DataRow(cells=cells)
                data_rows.append(data_row)
            self.content.controls[1].content.controls[2].content.columns = columns
            self.content.controls[1].content.controls[2].content.rows = data_rows

            self.page.dialog = self.alter_dialog_summary
            self.alter_dialog_summary.open = False
            latest_date_old_object = datetime.strptime(f"{latest_date}".replace("-",""), '%Y%m%d')
            latest_date_new_object = latest_date_old_object.strftime('%d%m%Y')
            self.number_of_version_kpe.content.value = f"{self.user_id}-{latest_date_new_object}-{latest_number}"
            self.page.update()
        except Exception as e:
            self.show_block_dialog("Ошибка при выборке данных", "Ошибка")
            print(f"Ошибка при выборке данных: {str(e)}")


    def close_dlg_summary(self, e):
        self.page.dialog = self.alter_dialog_summary
        self.alter_dialog_summary.open = False
        self.page.update()

    def form_report(self, e):
        selected_report_type = self.report_template.content.value
        if selected_report_type == "Карта КПЭ":
            self.report_depart.content.value = ''
            self.report_quater.content.value = ''
            self.report_spec.content.value = ''
            self.dropdown_options_specialists.clear()
            self.open_dialog_kpe()
        # * РАСЧЕТ ПРЕМИИ
        elif selected_report_type == "Расчет премии":
            self.report_depart.content.value = ''
            self.report_quater.content.value = ''
            self.report_spec.content.value = ''
            self.dropdown_options_specialists.clear()
            self.open_dialog()
        # *СВОДНЫЕ ДАННЫЕ ПО ИСПОЛНЕНИЮ
        elif selected_report_type == "Сводные данные по исполнению":
            self.report_depart.content.value = ''
            self.report_quater.content.value = ''
            self.report_spec.content.value = ''
            self.dropdown_options_specialists.clear()
            self.open_dialog_summary()
        else:
            self.show_block_dialog("Вы не выбрали шаблон отчета", "Ошибка")











    def export_report_to_excel(self, e):
        try:
            global results
            global results_summary
            global results_premi
            date = datetime.now()
            formatted_date = date.strftime("%d.%m.%Y %H.%M.%S")
            print(formatted_date)
            selected_report_type = self.report_template.content.value
            if selected_report_type == "Карта КПЭ":
                cursor = connection.cursor()
                cursor.execute(f"""
                SELECT sp.position, di.name
                FROM specialists AS sp
                    INNER JOIN name_of_department AS di ON di.department_id = sp.specialist_department_id
                WHERE sp.full_name = '{str(self.report_spec.content.value)}'
                """)
                position_name_dep = cursor.fetchone()
                
                cursor.execute(
                    f"SELECT specialist_id FROM specialists WHERE full_name='{str(self.report_spec.content.value)}';")
                specialist_id = cursor.fetchone()[0]
                
                cursor.execute(
                    f"""
                    SELECT 
                        concat(
                            toString(kpe_user_id), '-',
                            substring(replace(toString(date), '-', ''), 7, 2),
                            substring(replace(toString(date), '-', ''), 5, 2),
                            substring(replace(toString(date), '-', ''), 1, 4),
                            '-', toString(number)
                        ) AS number_of_version
                    FROM kpe_table
                    WHERE 
                        kpe_specialist_id = {int(specialist_id)} AND 
                        date = (
                            SELECT MAX(date)
                            FROM kpe_table
                            WHERE kpe_specialist_id = {int(specialist_id)}
                        ) AND 
                        number = (
                            SELECT MAX(number)
                            FROM kpe_table
                            WHERE 
                                kpe_specialist_id = {int(specialist_id)} AND 
                                date = (
                                    SELECT MAX(date)
                                    FROM kpe_table
                                    WHERE kpe_specialist_id = {int(specialist_id)}
                                )
                        );
                    """
                )
                latest_version = cursor.fetchone()[0]

                print(results)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                headers = ["п/п", "Наименование показателя", "Ед.изм.", "1 кв.", "2 кв.", "3 кв.", "4 кв.", "год",
                            "Вес КПЭ 1 кв.", "Вес КПЭ 2 кв.", "Вес КПЭ 3 кв.", "Вес КПЭ 4 кв."]
                sheet.column_dimensions["B"].width = 50
                
                arr_of_cols = ['C', 'D','E', 'F','G','I','J','K','L']
                for char in arr_of_cols:
                    sheet.column_dimensions[char].width = 15
                
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=8, column=col)
                    cell.value = header
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=14, name=main_font)

                # Merge cells D1 to I1 and add the specified phrase
                sheet.merge_cells('D1:I1')
                merged_cell = sheet['D1']
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text

                # Set the value for one of the constituent cells
                sheet['D1'].value = "УТВЕРЖДАЮ\nРуководитель агентства по труду и занятости населения Сахалинской области"
                
                sheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet['D1'].font = Font(size=14, name=main_font)

                # Repeat this pattern for other merged cells
                sheet.merge_cells('D2:I2')
                sheet['D2'].value = "______________Т.Г. Бабич"
                sheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D2'].font = Font(size=14, name=main_font)

                sheet.merge_cells('D3:I3')
                sheet['D3'].value = '"__" _________20__ года'
                sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D3'].font = Font(size=14, name=main_font)

                long_text = (
                    "КАРТА КЛЮЧЕВЫХ ПОКАЗАТЕЛЕЙ ЭФФЕКТИВНОСТИ ПРОФЕССИОНАЛЬНОЙ СЛУЖЕБНОЙ ДЕЯТЕЛЬНОСТИ "
                    "ГОСУДАРСТВЕННОГО ГРАЖДАНСКОГО СЛУЖАЩЕГО САХАЛИНСКОЙ ОБЛАСТИ"
                )

                # Merge the cells for the long text
                sheet.merge_cells('A4:I4')

                # Set the value for the merged cell (only set in the top-left cell of the merged range)
                sheet['A4'].value = long_text
                sheet['A4'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                sheet['A4'].font = Font(size=14, name=main_font, bold=True)

                sheet.merge_cells('A5:I5')
                sheet['A5'].value = 'Агентство по труду и занятости населения Сахалинской области'
                sheet['A5'].font = Font(bold=True, size=14, name=main_font)
                sheet['A5'].alignment = Alignment(horizontal='center', vertical='center')

                sheet.merge_cells('A6:I6')
                sheet['A6'].value = f'Карта КПЭ на {datetime.now().year} год {str(position_name_dep[0]).lower()}a {str(position_name_dep[1]).lower()} {str(self.report_spec.content.value)}'
                sheet['A6'].font = Font(bold=True, size=14, name=main_font)
                sheet['A6'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                for i, row_data in enumerate(results, start=9):  # Start the data from row 9
                    for col, value in enumerate(row_data, start=1):
                        cell = sheet.cell(row=i, column=col)
                        sheet.cell(row=i, column=col).value = value
                        sheet.cell(row=i, column=col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.font = Font(size=14, name=main_font)

                # Set alignment for the entire sheet
                last_row = sheet.max_row + 2
                
                
                parts = str(self.report_spec.content.value).split()
                # Берем первую букву фамилии и добавляем точку
                initials = parts[0] + " " + parts[1][0] + "." + parts[2][0] + "."
                
                name_of_dep= str(position_name_dep[1]).lower()
                
                # Заменяем "-ие" на "-ая"
                if "управление" in name_of_dep:
                    name_of_dep = name_of_dep.replace("управление", "управления")
                print(name_of_dep)  # Выведет "управлением обеспечения"

                position = str(position_name_dep[0])
                if "управления" in position:
                    position = position.replace("управления", "").strip()
                print(position)
                sheet.cell(row=last_row, column=1).value = "{} {} {} {}".format(position, name_of_dep,'_'*22, initials)

                
                sheet.merge_cells(f'A{last_row}:L{last_row}')  # Merge the cells for the record
                sheet.cell(row=last_row, column=1).font = Font(size=14, name=main_font)

                sheet.cell(row = last_row+1, column=10).value = f'Номер версии: {latest_version}'
                sheet.cell(row = last_row+1, column=10).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'J{last_row+1}:L{last_row+1}')
                
                sheet.row_dimensions[1].height = 50
                sheet.row_dimensions[4].height = 50
                sheet.row_dimensions[6].height = 30
                
                employee_folder = os.path.join(path, position_name_dep[1], self.report_spec.content.value)
                if not os.path.exists(employee_folder):
                    os.makedirs(employee_folder)

                filename = os.path.join(employee_folder, f"Карта КПЭ - {self.report_spec.content.value} - {formatted_date}.xlsx")
                print(filename)
                if filename:
                    workbook.save(filename)

                self.show_block_dialog("Отчет карта кпэ был выгружен в Excel", "Успешно")
            
    # * РАСЧЕТ ПРЕМИИ
            elif selected_report_type == "Расчет премии":
                cursor = connection.cursor()
                cursor.execute(
                    f"SELECT specialist_id FROM specialists WHERE full_name='{str(self.report_spec.content.value)}';")
                specialist_id = cursor.fetchone()[0]
                
                cursor.execute(
                    f"""
                    SELECT 
                        concat(
                            toString(kpe_user_id), '-',
                            substring(replace(toString(date), '-', ''), 7, 2),
                            substring(replace(toString(date), '-', ''), 5, 2),
                            substring(replace(toString(date), '-', ''), 1, 4),
                            '-', toString(number)
                        ) AS number_of_version
                    FROM kpe_table
                    WHERE 
                        kpe_specialist_id = {int(specialist_id)} AND 
                        date = (
                            SELECT MAX(date)
                            FROM kpe_table
                            WHERE kpe_specialist_id = {int(specialist_id)}
                        ) AND 
                        number = (
                            SELECT MAX(number)
                            FROM kpe_table
                            WHERE 
                                kpe_specialist_id = {int(specialist_id)} AND 
                                date = (
                                    SELECT MAX(date)
                                    FROM kpe_table
                                    WHERE kpe_specialist_id = {int(specialist_id)}
                                )
                        );
                    """
                )
                latest_version = cursor.fetchone()[0]
                
                cursor.execute(f"""
                SELECT sp.position, di.name
                FROM specialists AS sp
                    INNER JOIN name_of_department AS di ON di.department_id = sp.specialist_department_id
                WHERE sp.full_name = '{str(self.report_spec.content.value)}'
                """)
                position_name_dep = cursor.fetchone()
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                headers = ["№ пп.", 
                            "Наименование показателя",
                            "Единица измерения или срок",
                            "План",
                            "Факт",
                            "Вес КПЭ, %",
                            "Доля премии по факту выполнения показателя*",
                        ]

                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=13, column=col)
                    cell.value = header
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.font = Font(size=14, name=main_font)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                column_widths = {
                    'B': 45,
                    'C': 15,
                    'D': 35,
                    'E': 15,
                    'F': 25,
                    'G': 25
                }
                for column, width in column_widths.items():
                    sheet.column_dimensions[column].width = width

                sheet.merge_cells('D1:I1')
                merged_cell = sheet['D1']
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text

                # Set the value for one of the constituent cells
                sheet['D1'].value = "УТВЕРЖДАЮ\nРуководитель агентства по труду и занятости населения Сахалинской области"
                # sheet.row_dimensions[1].height = 40
                sheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet['D1'].font = Font(size=14, name=main_font)

                # Repeat this pattern for other merged cells
                sheet.merge_cells('D2:I2')
                sheet['D2'].value = "______________Т.Г. Бабич"
                sheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D2'].font = Font(size=14, name=main_font)

                sheet.merge_cells('D3:I3')
                sheet['D3'].value = '"__" _________20__ года'
                sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D3'].font = Font(size=14, name=main_font)
                
                texts = [
                    ('Расчет коэффициента выполнения', 4),
                    ('ключевых показателей эффективности', 5),
                    ('профессиональной служебной деятельности', 6),
                    ('по итогам работы за {} квартал'.format(self.report_quater.content.value), 7),
                    ('{}, {}'.format(position_name_dep[0], self.report_spec.content.value), 9),
                    ('(должность, ФИО гражданского служащего)', 10),
                    ('{}'.format(position_name_dep[1]), 11),
                    ('(наименование управления)', 12)
                ]

                # Проход по списку текстов и их строкам
                for text, row_num in texts:
                    merge_cell = 'B{}:F{}'.format(row_num, row_num)
                    sheet.merge_cells(merge_cell)
                    cell = sheet['B{}'.format(row_num)]
                    cell.value = text
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True, size=14, name=main_font)
                
                for col in range(1, 8):
                    sheet.cell(row=14, column=col).value = col
                    sheet.cell(row=14, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet.cell(row=14, column=col).font = Font(size=14, name=main_font)
                    cell = sheet.cell(row=14, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # цикл для заполнения таблицы
                for i, row_data in enumerate(results_premi, start=15):  # Start the data from row 9
                    for col, value in enumerate(row_data, start=1):
                        cell = sheet.cell(row=i, column=col)
                        sheet.cell(row=i, column=col).value = value
                        sheet.cell(row=i, column=col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        sheet.cell(row=i, column=col).font = Font(size=14, name=main_font)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    # sheet.row_dimensions[i].height = 75
                
                # текст расположенный снизу таблицы
                last_row = sheet.max_row
                sheet.cell(row=last_row+1, column=2).value = "Коэффициент выполнения КПЭ:"
                sheet.cell(row=last_row+1, column=2).alignment = Alignment(wrap_text=True)
                sheet.cell(row=last_row+1, column=2).font = Font(size=14, name=main_font)
                # sheet.row_dimensions[last_row+1].height = 25
                
                for col in range(1, 8):
                    cell = sheet.cell(row=last_row+1, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                sheet.cell(row=last_row+3, column=1).value = "_____________________________________________________"
                sheet.cell(row=last_row+3, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+3}:C{last_row+3}')
                
                sheet.cell(row=last_row+3, column=7).value = self.report_spec.content.value
                sheet.cell(row=last_row+3, column=7).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+3}:C{last_row+3}')
                
                sheet.cell(row=last_row+4, column=1).value = "_________________________________________"
                sheet.cell(row=last_row+4, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+4}:B{last_row+4}')
                
                sheet.cell(row=last_row+5, column=1).value = "(должность непосредственного руководителя гражданского служащего)"
                sheet.cell(row=last_row+5, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+5}:C{last_row+5}')
                # sheet.row_dimensions[last_row+5].height = 30
                sheet.cell(row=last_row+5, column=1).alignment = Alignment(wrap_text=True)
                
                sheet.cell(row=last_row+5, column=5).value = "(подпись)"
                sheet.cell(row=last_row+5, column=5).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+5, column=7).value = "(ФИО)"
                sheet.cell(row=last_row+5, column=7).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+7, column=1).value = "Ознакомлен:"
                sheet.cell(row=last_row+7, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+7}:C{last_row+7}')
                
                sheet.cell(row=last_row+8, column=1).value = "____________________________________"
                sheet.cell(row=last_row+8, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+7}:C{last_row+7}')
                
                sheet.cell(row=last_row+8, column=7).value = "___________________"
                sheet.cell(row=last_row+8, column=7).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+9, column=4).value = "(подпись)"
                sheet.cell(row=last_row+9, column=4).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+9, column=6).value = "(ФИО гражданского служащего)"
                sheet.cell(row=last_row+9, column=6).font = Font(size=14, name=main_font)
                sheet.cell(row=last_row+9, column=6).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                sheet.merge_cells(f'F{last_row+9}:G{last_row+9}')
                
                sheet.cell(row=last_row+10, column=1).value = "Согласовано:"
                sheet.cell(row=last_row+10, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+10}:C{last_row+10}')
                
                sheet.cell(row=last_row+11, column=1).value = "____________________________________"
                sheet.cell(row=last_row+11, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+11}:C{last_row+11}')
                
                sheet.cell(row=last_row+12, column=7).value = "___________________"
                sheet.cell(row=last_row+12, column=7).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+13, column=4).value = "(подпись)"
                sheet.cell(row=last_row+13, column=4).font = Font(size=14, name=main_font)
                
                sheet.cell(row=last_row+13, column=6).value = "(ФИО курирующего заместителя руководителя агентства)"
                sheet.cell(row=last_row+13, column=6).font = Font(size=14, name=main_font)
                sheet.cell(row=last_row+13, column=6).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # sheet.row_dimensions[last_row+13].height = 30
                sheet.merge_cells(f'F{last_row+13}:G{last_row+13}')
                
                sheet.cell(row=last_row+15, column=1).value = "_____________________________________________________________"
                sheet.cell(row=last_row+15, column=1).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'A{last_row+15}:D{last_row+15}')
                
                sheet.cell(row=last_row+16, column=2).value = "* - Гр. 7 = гр. 6/100 – если установленный показатель выполнен в полном объеме или перевыполнен.				"
                sheet.cell(row=last_row+16, column=2).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'B{last_row+16}:F{last_row+16}')
                
                sheet.cell(row=last_row+17, column=2).value = "- Гр. 7 = 0 – если установленный показатель не выполнен или выполнен не в полном объеме.»				"
                sheet.cell(row=last_row+17, column=2).font = Font(size=14, name=main_font)
                sheet.merge_cells(f'B{last_row+17}:F{last_row+17}')
                
                sheet[f'G{last_row+17}'].value = f'Номер версии: {latest_version}'
                sheet[f'G{last_row+17}'].font = Font(size=14, name=main_font)
                
                sheet.row_dimensions[1].height = 50
                sheet.row_dimensions[34].height = 30

                employee_folder = os.path.join(path, position_name_dep[1], self.report_spec.content.value)
                if not os.path.exists(employee_folder):
                    os.makedirs(employee_folder)

                filename = os.path.join(employee_folder, f"Расчет премии - {self.report_spec.content.value} - {formatted_date}.xlsx")
                print(filename)

                if filename:
                    workbook.save(filename)

                self.show_block_dialog("Отчет расчет приемии был выгружен в Excel", "Успешно")
    # *СВОДНЫЕ ДАННЫЕ ПО ИСПОЛНЕНИЮ
            elif selected_report_type == "Сводные данные по исполнению":
                cursor = connection.cursor()
                query_select = f'SELECT MAX(date), MAX(number) FROM kpe_table'
                cursor.execute(query_select)
                latest_date, latest_number = cursor.fetchone()
                latest_date_old_object = datetime.strptime(f"{latest_date}".replace("-",""), '%Y%m%d')
                latest_date_new_object = latest_date_old_object.strftime('%d%m%Y')
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                
                sheet.merge_cells('D1:F1')
                merged_cell = sheet['D1']
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text

                # Set the value for one of the constituent cells
                sheet['D1'].value = "УТВЕРЖДАЮ\nРуководитель агентства по труду и занятости населения Сахалинской области"
                # sheet.row_dimensions[1].height = 40
                sheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet['D1'].font = Font(size=14, name=main_font)

                # Repeat this pattern for other merged cells
                sheet.merge_cells('D2:F2')
                sheet['D2'].value = "______________Т.Г. Бабич"
                sheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D2'].font = Font(size=14, name=main_font)

                sheet.merge_cells('D3:F3')
                sheet['D3'].value = '"__" _________20__ года'
                sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['D3'].font = Font(size=14, name=main_font)
                sheet.merge_cells('A4:F4')

                sheet['A4'].value = f"РЕКОМЕНДУЕМАЯ ФОРМА СВЕДЕНИЙ О ВЫПОЛНЕНИИ КЛЮЧЕВЫХ ПОКАЗАТЕЛЕЙ ЭФФЕКТИВНОСТИ ПРОФЕССИОНАЛЬНОЙ СЛУЖЕБНОЙ ДЕЯТЕЛЬНОСТИ ГОСУДАРСТВЕННЫХ ГРАЖДАНСКИХ СЛУЖАЩИХ САХАЛИНСКОЙ ОБЛАСТИ\nза {self.report_quater.content.value} {datetime.now().year}г.\n(отчетный период)"
                sheet['A4'].font = Font(bold=True, size=14, name=main_font)
                sheet['A4'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                headers = ["№ п/п", 
                            "Наименование структурного подразделения Правительства Сахалинской области, государственного органа или органа исполнительной власти Сахалинской области",
                            "Наименование структурного подразделения органа исполнительной власти Сахалинской области",
                            "Должность",
                            "ФИО",
                            "Процент выполнения КПЭ по итогам отчетного периода, %"
                            ]

                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=5, column=col)
                    cell.value = header
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.font = Font(size=14, name=main_font)

                #Columns width
                arr_of_cols_1 = ['A','B','C','D', 'E', 'F']
                for cell in arr_of_cols_1:
                    sheet.column_dimensions[cell].width = 25
                
                for col in range(1, 7):
                    sheet.cell(row=6, column=col).value = col
                    sheet.cell(row=6, column=col).font = Font(italic=True)
                    sheet.cell(row=6, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell = sheet.cell(row=6, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.font = Font(size=14, name=main_font)

                for i, row_data in enumerate(results_summary, start=7):  # Start the data from row 9
                    for col, value in enumerate(row_data, start=1):
                        cell = sheet.cell(row=i, column=col)
                        sheet.cell(row=i, column=col).value = value
                        sheet.cell(row=i, column=col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.font = Font(size=14, name=main_font)
                    # sheet.row_dimensions[i].height = 34

                last_row = sheet.max_row
                sheet[f"F{last_row+1}"].value = f"Номер версии: {self.user_id}-{latest_date_new_object}-{latest_number}"
                sheet[f"F{last_row+1}"].font = Font(size=14, name=main_font)
                sheet[f"F{last_row+1}"].alignment = Alignment(horizontal='center', vertical='center')
                
                sheet.row_dimensions[1].height = 50
                sheet.row_dimensions[4].height = 50

                employee_folder = os.path.join(path, self.report_depart.content.value)
                if not os.path.exists(employee_folder):
                    os.makedirs(employee_folder)

                filename = os.path.join(employee_folder, f"Сводный отчет - {formatted_date}.xlsx")
                print(filename)

                if filename:
                    workbook.save(filename)

                self.show_block_dialog("Отчет сводных данных был выгружен в Excel", "Успешно")
            else:
                self.show_block_dialog("Вы не выбрали шаблон отчета для экспорта", "Ошибка")
        except TypeError:
            self.show_block_dialog("Вы не сформировали карту КПЭ", "Ошибка")
        except NameError:
            if self.report_template.content.value == "Сводные данные по исполнению":
                self.show_block_dialog("Вы не сформировали сводные данные по исполнению", "Ошибка")
            else:
                self.show_block_dialog("Вы не сформировали расчет премии", "Ошибка")
